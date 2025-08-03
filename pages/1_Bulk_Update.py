import streamlit as st
import pandas as pd
from io import BytesIO
import os
from openpyxl.styles import PatternFill

# Import Google libraries
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

st.set_page_config(
    page_title="Bulk Creative Updater",
    layout="wide"
)

st.title("Bulk Creative Updater Workflow")

# --- Tracker Type Maps ---
TRACKER_MAP_HOSTED_VIDEO = {
    "Impression": "THIRD_PARTY_URL_TYPE_IMPRESSION",
    "Click tracking": "THIRD_PARTY_URL_TYPE_CLICK_TRACKING",
    "Start": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_START",
    "First quartile": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_FIRST_QUARTILE",
    "Midpoint": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_MIDPOINT",
    "Third quartile": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_THIRD_QUARTILE",
    "Complete": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_COMPLETE",
}

# --- Authentication ---
SCOPES = ['https://www.googleapis.com/auth/display-video']

def get_creds():
    if 'creds' in st.session_state and st.session_state.creds and st.session_state.creds.valid:
        return st.session_state.creds
    if os.path.exists('token.json'):
        try:
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
            if creds and creds.valid:
                st.session_state.creds = creds
                return creds
        except Exception as e:
            st.warning(f"Could not load token.json: {e}. Please re-authenticate on the main page.")
            return None
    st.error("You are not logged in. Please go to the 'app.py' welcome page to authenticate.")
    return None

def fetch_creative_details(service, advertiser_id, creative_id):
    """Makes a live API call to fetch creative details."""
    try:
        request = service.advertisers().creatives().get(
            advertiserId=advertiser_id,
            creativeId=creative_id
        )
        return request.execute()
    except Exception as e:
        st.error(f"Failed to fetch Creative ID {creative_id}: {e}")
        return None

def generate_excel_file(df, is_report=False):
    """Generates a color-coded Excel file in memory."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Trackers')
        workbook = writer.book
        worksheet = writer.sheets['Trackers']
        
        # Define fills for coloring
        light_grey_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        if is_report:
            added_fill = PatternFill(start_color='D6EFD6', end_color='D6EFD6', fill_type='solid') # Green
            deleted_fill = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid') # Red
            updated_fill = PatternFill(start_color='D6E8EF', end_color='D6E8EF', fill_type='solid') # Blue

            for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
                status = getattr(row_data, 'status', '')
                fill = None
                if status == 'ADDED':
                    fill = added_fill
                elif status == 'DELETED':
                    fill = deleted_fill
                elif status == 'UPDATED':
                    fill = updated_fill
                
                if fill:
                    for col_num in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_num, column=col_num).fill = fill
        else: # Standard alternating colors
            current_creative_id = None
            use_grey = False
            for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
                if str(row_data.creative_id) != str(current_creative_id):
                    current_creative_id = str(row_data.creative_id)
                    use_grey = not use_grey
                if use_grey:
                    for col_num in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_num, column=col_num).fill = light_grey_fill

    return output.getvalue()

# --- Main UI ---
creds = get_creds()

if creds:
    # --- Initialize Session State ---
    if 'processed_df' not in st.session_state:
        st.session_state.processed_df = None
    if 'individual_results' not in st.session_state:
        st.session_state.individual_results = None
    if 'update_plan' not in st.session_state:
        st.session_state.update_plan = None
    if 'change_report_updates' not in st.session_state:
        st.session_state.change_report_updates = None
    if 'change_report_adds_deletes' not in st.session_state:
        st.session_state.change_report_adds_deletes = None


    # --- Phase 1: Uploader ---
    st.header("Phase 1: Upload Creative IDs")
    advertiser_id_input = st.text_input("Enter the Advertiser ID for all creatives")
    uploaded_ids_file = st.file_uploader("Upload a one-column CSV with your Creative IDs", type="csv")

    if st.button("Process IDs and Show Results"):
        if uploaded_ids_file and advertiser_id_input:
            try:
                raw_text = uploaded_ids_file.getvalue().decode('utf-8')
                lines = raw_text.splitlines()
                creative_ids = [line.strip() for line in lines if line.strip() and line.strip().lower() != 'creative_id']

                if not creative_ids:
                    st.error("The uploaded file contains no valid Creative IDs.")
                else:
                    service = build('displayvideo', 'v3', credentials=creds)
                    all_trackers_data = []
                    individual_results_list = []
                    
                    with st.spinner(f"Fetching data for {len(creative_ids)} creatives..."):
                        progress_bar = st.progress(0)
                        for i, creative_id in enumerate(creative_ids):
                            details = fetch_creative_details(service, advertiser_id_input, creative_id)
                            individual_results_list.append(details)
                            
                            if details:
                                trackers = details.get("thirdPartyUrls", [])
                                creative_name = details.get("displayName", "N/A")
                                reverse_map = {v: k for k, v in TRACKER_MAP_HOSTED_VIDEO.items()}

                                if trackers:
                                    for tracker in trackers:
                                        api_type = tracker.get('type')
                                        event_type = reverse_map.get(api_type, api_type)
                                        all_trackers_data.append({
                                            "advertiser_id": advertiser_id_input,
                                            "creative_id": creative_id,
                                            "creative_name": creative_name,
                                            "event_type": event_type,
                                            "existing_url": tracker.get("url"),
                                            "new_url": ""
                                        })
                                else:
                                    all_trackers_data.append({
                                        "advertiser_id": advertiser_id_input,
                                        "creative_id": creative_id,
                                        "creative_name": creative_name,
                                        "event_type": "",
                                        "existing_url": "",
                                        "new_url": ""
                                    })
                            progress_bar.progress((i + 1) / len(creative_ids))
                    
                    st.session_state.individual_results = individual_results_list
                    st.session_state.processed_df = pd.DataFrame(all_trackers_data)
                    st.success("Data extraction complete.")
            except Exception as e:
                st.error(f"An error occurred: {e}")
        else:
            st.warning("Please provide an Advertiser ID and upload a file.")

    # --- Display Results and Global Download Button ---
    if st.session_state.individual_results:
        st.header("Extracted Creative Details")
        for creative_data in st.session_state.individual_results:
            if creative_data:
                name = creative_data.get('displayName', 'N/A')
                c_id = creative_data.get('creativeId', 'N/A')
                with st.expander(f"Creative: {name} (ID: {c_id})"):
                    # Display a clean table inside the expander
                    trackers = creative_data.get("thirdPartyUrls", [])
                    if trackers:
                        reverse_map = {v: k for k, v in TRACKER_MAP_HOSTED_VIDEO.items()}
                        display_data = [{"event_type": reverse_map.get(t.get('type'), t.get('type')), "url": t.get('url')} for t in trackers]
                        st.dataframe(pd.DataFrame(display_data))
                    else:
                        st.write("No third-party trackers found.")

    if st.session_state.processed_df is not None and not st.session_state.processed_df.empty:
        st.header("Download Combined File")
        excel_data = generate_excel_file(st.session_state.processed_df)
        st.download_button(
            label="📥 Download Combined Excel File to Edit",
            data=excel_data,
            file_name="dv360_trackers_to_edit.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    # --- Phase 2: Upload Edited File for Validation and Review ---
    st.header("Phase 2: Upload Your Edited Excel File")
    edited_file = st.file_uploader("Upload the Excel file you edited", type="xlsx")

    if edited_file:
        if st.button("Validate and Review Changes"):
            try:
                with st.spinner("Validating file and comparing changes..."):
                    edited_df = pd.read_excel(edited_file).fillna('')
                    original_df = st.session_state.processed_df.fillna('')

                    # Create unique keys for comparison
                    original_df['key'] = original_df['creative_id'].astype(str) + "|" + original_df['event_type']
                    edited_df['key'] = edited_df['creative_id'].astype(str) + "|" + edited_df['event_type']
                    
                    # --- NEW VALIDATION LOGIC ---
                    updates_list = []
                    adds_deletes_list = []
                    
                    # Find additions and deletions
                    original_keys = set(original_df['key'])
                    edited_keys = set(edited_df['key'])
                    
                    deleted_keys = original_keys - edited_keys
                    added_keys = edited_keys - original_keys
                    
                    if deleted_keys:
                        deleted_df = original_df[original_df['key'].isin(deleted_keys)].copy()
                        deleted_df['status'] = 'DELETED'
                        adds_deletes_list.append(deleted_df)

                    if added_keys:
                        added_df = edited_df[edited_df['key'].isin(added_keys)].copy()
                        added_df['status'] = 'ADDED'
                        adds_deletes_list.append(added_df)

                    # Find updates
                    common_keys = original_keys.intersection(edited_keys)
                    for key in common_keys:
                        original_url = original_df[original_df['key'] == key]['existing_url'].iloc[0]
                        edited_new_url = edited_df[edited_df['key'] == key]['new_url'].iloc[0]
                        if edited_new_url and edited_new_url != original_url:
                            update_row = edited_df[edited_df['key'] == key].copy()
                            update_row['status'] = 'UPDATED'
                            updates_list.append(update_row)
                    
                    st.subheader("Validation Complete")
                    if not adds_deletes_list and not updates_list:
                        st.success("✅ No changes were detected in the uploaded file.")
                        st.session_state.update_plan = None # Prevent Phase 3 from showing
                    else:
                        st.info("Please review the changes below by downloading the reports.")
                        st.session_state.update_plan = edited_df # Allow Phase 3 to proceed

                    # Store reports in session state for download
                    st.session_state.change_report_updates = pd.concat(updates_list) if updates_list else None
                    st.session_state.change_report_adds_deletes = pd.concat(adds_deletes_list) if adds_deletes_list else None

            except Exception as e:
                st.error(f"An error occurred during validation: {e}")

    # --- Download Change Reports ---
    if st.session_state.get('change_report_updates') is not None:
        report_excel = generate_excel_file(st.session_state.change_report_updates, is_report=True)
        st.download_button(
            label="📊 Download Update Report",
            data=report_excel,
            file_name="update_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    if st.session_state.get('change_report_adds_deletes') is not None:
        report_excel = generate_excel_file(st.session_state.change_report_adds_deletes, is_report=True)
        st.download_button(
            label="🚨 Download Add/Delete Report",
            data=report_excel,
            file_name="add_delete_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    # --- Phase 3: Final Confirmation ---
    if st.session_state.get('update_plan') is not None:
        st.header("Phase 3: Confirm and Push to DV360")
        st.warning("⚠️ **FINAL WARNING:** This action is irreversible.")
        
        if st.button("Confirm and Send to DV360", type="primary"):
            try:
                with st.spinner("Sending updates to the DV360 API..."):
                    plan_df = st.session_state.update_plan
                    service = build('displayvideo', 'v3', credentials=creds)

                    for creative_id, group in plan_df.groupby('creative_id'):
                        final_trackers = []
                        adv_id = group['advertiser_id'].iloc[0]
                        for _, row in group.iterrows():
                            url_to_use = row['new_url'] if pd.notna(row['new_url']) and str(row['new_url']).strip() else row['existing_url']
                            if pd.notna(row['event_type']) and pd.notna(url_to_use):
                                api_type = TRACKER_MAP_HOSTED_VIDEO.get(row['event_type'], row['event_type'])
                                final_trackers.append({"type": api_type, "url": str(url_to_use)})
                        
                        service.advertisers().creatives().patch(
                            advertiserId=str(adv_id),
                            creativeId=str(creative_id),
                            updateMask="thirdPartyUrls",
                            body={"thirdPartyUrls": final_trackers}
                        ).execute()

                    st.success("All updates have been processed successfully!")
                    # Clear session state to reset the app
                    for key in ['processed_df', 'individual_results', 'update_plan', 'change_report_updates', 'change_report_adds_deletes']:
                        if key in st.session_state:
                            del st.session_state[key]

            except Exception as e:
                st.error(f"An error occurred during the final update: {e}")

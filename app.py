import streamlit as st
import pandas as pd
from io import BytesIO
import os
from openpyxl.styles import PatternFill

# Import Google libraries
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

st.set_page_config(
    page_title="Bulk Creative Updater",
    layout="wide"
)

st.title("Bulk Creative Updater Workflow")

# --- Tracker Type Maps ---
TRACKER_MAP_HOSTED_VIDEO = {
    "Impression": "THIRD_PARTY_URL_TYPE_IMPRESSION",
    "Click tracking": "THIRD_PARTY_URL_TYPE_CLICK_TRACKING",
    "Start": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_START",
    "First quartile": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_FIRST_QUARTILE",
    "Midpoint": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_MIDPOINT",
    "Third quartile": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_THIRD_QUARTILE",
    "Complete": "THIRD_PARTY_URL_TYPE_AUDIO_VIDEO_COMPLETE",
}
# Add other maps (VAST, Standard) and a detection function if needed for other creative types

# --- Authentication ---
SCOPES = ['https://www.googleapis.com/auth/display-video']

def get_creds():
    if 'creds' in st.session_state and st.session_state.creds and st.session_state.creds.valid:
        return st.session_state.creds
    if os.path.exists('token.json'):
        try:
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
            if creds and creds.valid:
                st.session_state.creds = creds
                return creds
        except Exception as e:
            st.warning(f"Could not load token.json: {e}. Please re-authenticate on the main page.")
            return None
    st.error("You are not logged in. Please go to the 'app.py' welcome page to authenticate.")
    return None

def fetch_creative_details(service, advertiser_id, creative_id):
    """Makes a live API call to fetch creative details."""
    try:
        request = service.advertisers().creatives().get(
            advertiserId=advertiser_id,
            creativeId=creative_id
        )
        return request.execute()
    except Exception as e:
        st.error(f"Failed to fetch Creative ID {creative_id}: {e}")
        return None

def generate_excel_file(df):
    """Generates a color-coded Excel file in memory."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Trackers')
        workbook = writer.book
        worksheet = writer.sheets['Trackers']

        light_grey_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        current_creative_id = None
        use_grey = False

        for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
            if str(row_data.creative_id) != str(current_creative_id):
                current_creative_id = str(row_data.creative_id)
                use_grey = not use_grey

            if use_grey:
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = light_grey_fill
    
    return output.getvalue()


# --- Main UI ---
creds = get_creds()

if creds:
    st.header("Phase 1: Upload Creative IDs")
    advertiser_id = st.text_input("Enter the Advertiser ID for all creatives")
    uploaded_ids_file = st.file_uploader("Upload a one-column CSV with your Creative IDs", type="csv")

    if uploaded_ids_file and advertiser_id:
        if st.button("Process IDs and Show Results"):
            try:
                raw_text = uploaded_ids_file.getvalue().decode('utf-8')
                lines = raw_text.splitlines()
                creative_ids = [line.strip() for line in lines if line.strip() and line.strip().lower() != 'creative_id']

                if not creative_ids:
                    st.error("The uploaded file contains no valid Creative IDs.")
                else:
                    st.session_state.all_creative_data = [] # Store all data for combined download
                    st.session_state.individual_results = [] # Store individual results for display
                    service = build('displayvideo', 'v3', credentials=creds)
                    
                    with st.spinner(f"Fetching data for {len(creative_ids)} creatives..."):
                        progress_bar = st.progress(0)
                        for i, creative_id in enumerate(creative_ids):
                            details = fetch_creative_details(service, advertiser_id, creative_id)
                            st.session_state.individual_results.append(details) # Add for display
                            
                            # Process and add to the combined list for download
                            if details:
                                if details.get("thirdPartyUrls"):
                                    for tracker in details["thirdPartyUrls"]:
                                        st.session_state.all_creative_data.append({
                                            "creative_id": creative_id,
                                            "creative_name": details.get("displayName", "N/A"),
                                            "event_type": tracker.get("type"), # Store raw type for now
                                            "url": tracker.get("url")
                                        })
                                else:
                                    st.session_state.all_creative_data.append({
                                        "creative_id": creative_id,
                                        "creative_name": details.get("displayName", "N/A"),
                                        "event_type": "",
                                        "url": ""
                                    })
                            progress_bar.progress((i + 1) / len(creative_ids))
                    
                    st.success("Data extraction complete. Results are shown below.")

            except Exception as e:
                st.error(f"An error occurred: {e}")

    # --- Display Folded Results AND the Combined Download Button ---
    if 'individual_results' in st.session_state and st.session_state.individual_results:
        st.header("Extracted Creative Details")
        st.info("Click on each creative to view its trackers. Use the button below to download a combined Excel file.")

        # Display individual expanders
        for creative_data in st.session_state.individual_results:
            if creative_data:
                name = creative_data.get('displayName', 'N/A')
                c_id = creative_data.get('creativeId', 'N/A')
                with st.expander(f"Creative: {name} (ID: {c_id})"):
                    st.write(creative_data.get("thirdPartyUrls", "No third-party trackers found."))

        # Create and show the combined download button
        if st.session_state.all_creative_data:
            df_for_export = pd.DataFrame(st.session_state.all_creative_data)
            
            # Map the raw event types to friendly names for the export file
            reverse_map = {v: k for k, v in TRACKER_MAP_HOSTED_VIDEO.items()} # Assumes hosted video
            df_for_export['event_type'] = df_for_export['event_type'].map(reverse_map).fillna(df_for_export['event_type'])

            excel_data = generate_excel_file(df_for_export)
            st.download_button(
                label="📥 Download Combined Excel File to Edit",
                data=excel_data,
                file_name="dv360_trackers_to_edit.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # --- Placeholder for Phase 2 and 3 ---
    st.header("Phase 2: Upload Your Edited Excel File")
    edited_file = st.file_uploader("Upload the Excel file you edited", type="xlsx")

    if edited_file:
        if st.button("Validate and Review Changes"):
            st.info("Validation and review step would be implemented here.")
